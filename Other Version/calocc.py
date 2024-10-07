import pandas as pd
import heapq
from openpyxl import load_workbook

# Read the Excel file
dataframe = pd.read_excel('datanode2.xlsx')

# Check if 'Cost' column has only numerical values
def check_cost_column(dataframe):
    non_numerical_rows = dataframe[~dataframe['Cost'].apply(lambda x: isinstance(x, (int, float)))]
    if not non_numerical_rows.empty:
        print("Error: 'Cost' column contains non-numerical values in the following rows:")
        print(non_numerical_rows)
        return False
    return True

# Validate the 'Cost' column
if not check_cost_column(dataframe):
    raise ValueError("The 'Cost' column contains non-numerical values. Please correct the data and try again.")

# Create a graph from the dataframe
graph = {}
nodes = set()
for index, row in dataframe.iterrows():
    start, end, cost, occupancy = row['Node Awal'], row['Node Tujuan'], row['Cost'], row['Occupancy']
    nodes.update([start, end])
    if start not in graph:
        graph[start] = []
    if end not in graph:
        graph[end] = []
    graph[start].append((cost, end, 0, occupancy))  # Initialize bandwidth to 0
    graph[end].append((cost, start, 0, occupancy))  # If the graph is undirected


def dijkstra(graph, start):
    # Priority queue to store (cost, node)
    queue = [(0, start)]
    # Dictionary to store the shortest path to each node
    shortest_paths = {start: (None, 0, 0)}  # (previous node, cost, bandwidth)
    while queue:
        (cost, node) = heapq.heappop(queue)
        for edge_cost, neighbor, bandwidth, occupancy in graph[node]:
            old_cost = shortest_paths.get(neighbor, (None, float('inf'), 0))[1]
            new_cost = cost + edge_cost
            if new_cost < old_cost:
                shortest_paths[neighbor] = (node, new_cost, bandwidth)
                heapq.heappush(queue, (new_cost, neighbor))
    return shortest_paths

# Function to reconstruct the shortest path
def reconstruct_path(shortest_paths, start, end):
    path = []
    while end is not None:
        path.append(end)
        end = shortest_paths[end][0]
    path.reverse()
    return path

# Function to update bandwidth and occupancy
def update_bandwidth(graph, path, new_bandwidth, dataframe, workbook):
    updated_info = []
    sheet = workbook.active
    for i in range(len(path) - 1):
        start, end = path[i], path[i + 1]
        for j in range(len(graph[start])):
            if graph[start][j][1] == end:
                cost, _, bandwidth, occupancy = graph[start][j]
                new_occupancy = occupancy + new_bandwidth
                graph[start][j] = (cost, end,  new_bandwidth, new_occupancy)
                updated_info.append((start, end,  new_bandwidth, new_occupancy))
                # Update the dataframe
                dataframe.loc[((dataframe['Node Awal'] == start) & (dataframe['Node Tujuan'] == end)) |
                              ((dataframe['Node Awal'] == end) & (dataframe['Node Tujuan'] == start)), 'Occupancy'] = new_occupancy
                # Update the Excel sheet directly
                for row in sheet.iter_rows(min_row=2, values_only=False):
                    if (row[0].value == start and row[1].value == end) or (row[0].value == end and row[1].value == start):
                        row[4].value = new_occupancy
        for j in range(len(graph[end])):
            if graph[end][j][1] == start:
                cost, _, bandwidth, occupancy = graph[end][j]
                new_occupancy = occupancy + new_bandwidth
                graph[end][j] = (cost, start, bandwidth + new_bandwidth, new_occupancy)
    return updated_info

# Function to save results to the second sheet in the existing Excel file
def save_results_to_excel(file_path, results):
    workbook = load_workbook(file_path)
    if 'Results2' not in workbook.sheetnames:
        sheet = workbook.create_sheet('Results2')
    else:
        sheet = workbook['Results2']
    for result in results:
        sheet.append(result)
    workbook.save(file_path)

# Manually input data
inputs = []
while True:
    user_id = input("Enter user ID (or 'done' to finish): ")
    if user_id.lower() == 'done':
        break
    start_node = input("Enter start node: ")
    end_node = input("Enter end node: ")
    new_bandwidth = float(input("Enter new bandwidth: "))
    inputs.append({"user_id": user_id, "start_node": start_node, "end_node": end_node, "new_bandwidth": new_bandwidth})

results = []
workbook = load_workbook('datanode2.xlsx')
for input_data in inputs:
    user_id = input_data["user_id"]
    start_node = input_data["start_node"]
    end_node = input_data["end_node"]
    new_bandwidth = input_data["new_bandwidth"]

    shortest_paths = dijkstra(graph, start_node)

    # Print shortest path and cost
    if end_node in shortest_paths:
        path = reconstruct_path(shortest_paths, start_node, end_node)
        cost = shortest_paths[end_node][1]
        updated_info = update_bandwidth(graph, path, new_bandwidth, dataframe, workbook)
        result = [user_id, ' -> '.join(path), cost]
        for info in updated_info:
            start, end, bandwidth, occupancy = info
            result.extend([f"{start}->{end}", bandwidth, occupancy])
        results.append(result)
        print(f"User {user_id}: Shortest path from {start_node} to {end_node} is: {' -> '.join(path)} with cost {cost}")
        print("Updated bandwidth and occupancy on each path:")
        for info in updated_info:
            start, end, bandwidth, occupancy = info
            print(f"Path {start} -> {end}: Bandwidth = {bandwidth}, New Occupancy = {occupancy}")
    else:
        print(f"User {user_id}: No path found from {start_node} to {end_node}")


# Save the updated workbook back to the Excel file
workbook.save('datanode2.xlsx')
# Save results to the Excel file
save_results_to_excel('datanode2.xlsx', results)