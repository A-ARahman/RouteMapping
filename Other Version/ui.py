import pandas as pd
import heapq
from openpyxl import load_workbook
import tkinter as tk
from tkinter import ttk, messagebox

# Read the Excel file
dataframe = pd.read_excel('datanode2.xlsx')

# Check if 'Cost' column has only numerical values
def check_cost_column(dataframe):
    non_numerical_rows = dataframe[~dataframe['Cost'].apply(lambda x: isinstance(x, (int, float)))]
    if not non_numerical_rows.empty:
        print("Error: 'Cost' column contains non-numerical values in the following rows:")
        print(non_numerical_rows)
        return False
    return True

# Validate the 'Cost' column
if not check_cost_column(dataframe):
    raise ValueError("The 'Cost' column contains non-numerical values. Please correct the data and try again.")

# Create a graph from the dataframe
graph = {}
nodes = set()
for index, row in dataframe.iterrows():
    start, end, cost, occupancy = row['Node Awal'], row['Node Tujuan'], row['Cost'], row['Occupancy']
    nodes.update([start, end])
    if start not in graph:
        graph[start] = []
    if end not in graph:
        graph[end] = []
    graph[start].append((cost, end, 0, occupancy))  # Initialize bandwidth to 0
    graph[end].append((cost, start, 0, occupancy))  # If the graph is undirected

def dijkstra(graph, start):
    queue = [(0, start)]
    shortest_paths = {start: (None, 0, 0)}  # (previous node, cost, bandwidth)
    while queue:
        (cost, node) = heapq.heappop(queue)
        for edge_cost, neighbor, bandwidth, occupancy in graph[node]:
            old_cost = shortest_paths.get(neighbor, (None, float('inf'), 0))[1]
            new_cost = cost + edge_cost
            if new_cost < old_cost:
                shortest_paths[neighbor] = (node, new_cost, bandwidth)
                heapq.heappush(queue, (new_cost, neighbor))
    return shortest_paths

def reconstruct_path(shortest_paths, start, end):
    path = []
    while end is not None:
        path.append(end)
        end = shortest_paths[end][0]
    path.reverse()
    return path

def update_bandwidth(graph, path, new_bandwidth, dataframe, workbook):
    updated_info = []
    sheet = workbook.active
    for i in range(len(path) - 1):
        start, end = path[i], path[i + 1]
        for j in range(len(graph[start])):
            if graph[start][j][1] == end:
                cost, _, bandwidth, occupancy = graph[start][j]
                new_occupancy = occupancy + new_bandwidth
                graph[start][j] = (cost, end,  new_bandwidth, new_occupancy)
                updated_info.append((start, end,  new_bandwidth, new_occupancy))
                dataframe.loc[((dataframe['Node Awal'] == start) & (dataframe['Node Tujuan'] == end)) |
                              ((dataframe['Node Awal'] == end) & (dataframe['Node Tujuan'] == start)), 'Occupancy'] = new_occupancy
                for row in sheet.iter_rows(min_row=2, values_only=False):
                    if (row[0].value == start and row[1].value == end) or (row[0].value == end and row[1].value == start):
                        row[4].value = new_occupancy
        for j in range(len(graph[end])):
            if graph[end][j][1] == start:
                cost, _, bandwidth, occupancy = graph[end][j]
                new_occupancy = occupancy + new_bandwidth
                graph[end][j] = (cost, start, bandwidth + new_bandwidth, new_occupancy)
    return updated_info

def save_results_to_excel(file_path, results):
    workbook = load_workbook(file_path)
    if 'Results2' not in workbook.sheetnames:
        sheet = workbook.create_sheet('Results2')
    else:
        sheet = workbook['Results2']
    for result in results:
        sheet.append(result)
    workbook.save(file_path)

def process_input():
    user_id = user_id_entry.get()
    start_node = start_node_var.get()
    end_node = end_node_var.get()
    try:
        new_bandwidth = float(new_bandwidth_entry.get())
    except ValueError:
        messagebox.showerror("Invalid Input", "Please enter a valid number for bandwidth.")
        return

    if not user_id or not start_node or not end_node:
        messagebox.showerror("Invalid Input", "Please fill in all fields.")
        return

    shortest_paths = dijkstra(graph, start_node)
    if end_node in shortest_paths:
        path = reconstruct_path(shortest_paths, start_node, end_node)
        cost = shortest_paths[end_node][1]
        updated_info = update_bandwidth(graph, path, new_bandwidth, dataframe, workbook)
        result = [user_id, ' -> '.join(path), cost]
        for info in updated_info:
            start, end, bandwidth, occupancy = info
            result.extend([f"{start}->{end}", bandwidth, occupancy])
        results.append(result)
        messagebox.showinfo("Result", f"User {user_id}: Shortest path from {start_node} to {end_node} is: {' -> '.join(path)} with cost {cost}")
    else:
        messagebox.showerror("No Path Found", f"User {user_id}: No path found from {start_node} to {end_node}")

    workbook.save('datanode2.xlsx')
    save_results_to_excel('datanode2.xlsx', results)

# Create the main window
root = tk.Tk()
root.title("Network Path Finder")

# Create and place the widgets
tk.Label(root, text="User ID:").grid(row=0, column=0, padx=10, pady=5)
user_id_entry = tk.Entry(root)
user_id_entry.grid(row=0, column=1, padx=10, pady=5)

tk.Label(root, text="Start Node:").grid(row=1, column=0, padx=10, pady=5)
start_node_var = tk.StringVar()
start_node_menu = ttk.Combobox(root, textvariable=start_node_var, values=list(nodes))
start_node_menu.grid(row=1, column=1, padx=10, pady=5)

tk.Label(root, text="End Node:").grid(row=2, column=0, padx=10, pady=5)
end_node_var = tk.StringVar()
end_node_menu = ttk.Combobox(root, textvariable=end_node_var, values=list(nodes))
end_node_menu.grid(row=2, column=1, padx=10, pady=5)

tk.Label(root, text="New Bandwidth:").grid(row=3, column=0, padx=10, pady=5)
new_bandwidth_entry = tk.Entry(root)
new_bandwidth_entry.grid(row=3, column=1, padx=10, pady=5)

submit_button = tk.Button(root, text="Submit", command=process_input)
submit_button.grid(row=4, column=0, columnspan=2, pady=10)

results = []
workbook = load_workbook('datanode2.xlsx')

# Start the Tkinter event loop
root.mainloop()