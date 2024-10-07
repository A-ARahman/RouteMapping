import pandas as pd
import heapq
from openpyxl import load_workbook
import time

# Function to convert 'NODE ASAL' and 'NODE TUJUAN' to strings
def convert_nodes_to_string(dataframe):
    dataframe['NODE ASAL'] = dataframe['NODE ASAL'].astype(str)
    dataframe['NODE TUJUAN'] = dataframe['NODE TUJUAN'].astype(str)

# Read the Excel file
dataframe = pd.read_excel('DIJKSTRA2.xlsx', sheet_name='TRUNK')

# Convert 'NODE ASAL' and 'NODE TUJUAN' to strings
convert_nodes_to_string(dataframe)

# Create a graph from the dataframe
graph = {}
for index, row in dataframe.iterrows():
    start, end, cost, status = row['NODE ASAL'], row['NODE TUJUAN'], row['COST'], row['STATUS_TRUNK']
    if start not in graph:
        graph[start] = []
    if end not in graph:
        graph[end] = []
    graph[start].append((cost, end, status))
    graph[end].append((cost, start, status))  # If the graph is undirected

def DIJKSTRA2(graph, start):
    # Priority queue to store (cost, node)
    queue = [(0, start)]
    # Dictionary to store the shortest path to each node
    shortest_paths = {start: (None, 0)}  # (previous node, cost)
    while queue:
        (cost, node) = heapq.heappop(queue)
        for edge_cost, neighbor, status in graph[node]:
            if status == 'DISABLED':
                continue  # Skip this edge if it is disabled
            old_cost = shortest_paths.get(neighbor, (None, float('inf')))[1]
            new_cost = cost + edge_cost
            if new_cost < old_cost:
                shortest_paths[neighbor] = (node, new_cost)
                heapq.heappush(queue, (new_cost, neighbor))
    return shortest_paths

# Function to reconstruct the shortest path
def reconstruct_path(shortest_paths, end):
    path = []
    while end is not None:
        path.append(end)
        end = shortest_paths[end][0]
    path.reverse()
    return path

def update_bandwidth_and_write_header(dataframe, path, user_id, bandwidth, workbook):
    # Add a new column for the user_id if it doesn't exist
    if user_id not in dataframe.columns:
        dataframe[user_id] = 0
    else:
        # Reset the values for the added column
        dataframe[user_id] = 0

    # Update the bandwidth for each edge in the path
    for i in range(len(path) - 1):
        start, end = path[i], path[i + 1]
        for j in range(len(dataframe)):
            if (dataframe.at[j, 'NODE ASAL'] == start and dataframe.at[j, 'NODE TUJUAN'] == end) or \
               (dataframe.at[j, 'NODE ASAL'] == end and dataframe.at[j, 'NODE TUJUAN'] == start):
                dataframe.at[j, user_id] += bandwidth

    # Write user ID as header
    sheet = workbook['TRUNK']
    # Check if the user ID header already exists
    if user_id not in [cell.value for cell in sheet[1]]:
        # Find the first empty column
        col_idx = len(sheet[1]) + 1
        sheet.cell(row=1, column=col_idx, value=user_id)

def remove_user_id_column(dataframe, input_data, workbook):
    # Get the list of user IDs from the input data
    input_user_ids = set(input_data['OPTY_ID'])
    
    # Get the list of user IDs from the TRUNK sheet
    sheet = workbook['TRUNK']
    trunk_user_ids = [cell.value for cell in sheet[1] if cell.value not in ['NODE ASAL', 'NODE TUJUAN', 'COST', 'KAPASITAS', 'PEAK UTILITAS','STATUS_TRUNK','LACK', 'TOTAL_BW_OPPTY']]
    
    # Find user IDs that are in TRUNK but not in input data
    user_ids_to_remove = set(trunk_user_ids) - input_user_ids
    
    for user_id in user_ids_to_remove:
        # Find the column index of the user ID
        col_idx = None
        for idx, cell in enumerate(sheet[1]):
            if cell.value == user_id:
                col_idx = idx + 1
                break
        
        if col_idx:
            # Delete the column
            sheet.delete_cols(col_idx)
            # Also remove the column from the dataframe
            if user_id in dataframe.columns:
                dataframe.drop(columns=[user_id], inplace=True)

def main():
    start_time = time.time()  # Start the timer

    # Read input data from Excel file
    input_data = pd.read_excel('DIJKSTRA2.xlsx', sheet_name='OPPTY')

    # Load the existing workbook
    workbook = load_workbook('DIJKSTRA2.xlsx')

    for index, row in input_data.iterrows():
        user_id = row['OPTY_ID']
        start_node = row['NODE_ASAL']
        end_node = row['NODE_TUJUAN']
        bandwidth = row['BW']
        shortest_paths = DIJKSTRA2(graph, start_node)
        path = reconstruct_path(shortest_paths, end_node)
        
        # Update the bandwidth in the TRUNK sheet and write user ID header
        update_bandwidth_and_write_header(dataframe, path, user_id, bandwidth, workbook)

    sheet = workbook['TRUNK']

    # Update only the necessary cells
    for index, row in dataframe.iterrows():
        for col in dataframe.columns:
            if col not in ['LACK', 'TOTAL_BW_OPPTY']:  # Replace with actual column names
                sheet.cell(row=index+2, column=dataframe.columns.get_loc(col)+1, value=row[col])

    # Recheck and remove user ID columns if necessary
    remove_user_id_column(dataframe, input_data, workbook)

    # Save the workbook
    workbook.save('DIJKSTRA2.xlsx')

    end_time = time.time()  # End the timer
    print(f"Execution time: {end_time - start_time} seconds")  # Print the execution time

if __name__ == "__main__":
    main()
